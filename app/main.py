
import os, json, datetime as dt, io, logging
from typing import List, Optional, Dict
from fastapi import FastAPI, Request, Depends, Form, HTTPException, Query
from fastapi.responses import RedirectResponse, StreamingResponse, PlainTextResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.interval import IntervalTrigger
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .db import init_db, SessionLocal
from .models import Booking, Staff, Apartment, Task, TimeLog
from .services_smoobu import SmoobuClient
from .utils import new_token, today_iso, now_iso
from .sync import upsert_tasks_from_bookings

ADMIN_TOKEN = os.getenv("ADMIN_TOKEN", "")
TIMEZONE = os.getenv("TIMEZONE", "Europe/Berlin")
REFRESH_INTERVAL_MINUTES = int(os.getenv("REFRESH_INTERVAL_MINUTES", "60"))
BASE_URL = os.getenv("BASE_URL", "")

log = logging.getLogger("smoobu")
logging.basicConfig(level=logging.INFO)

app = FastAPI(title="Smoobu Staff Planner Pro (v6.4)")

if os.path.isdir("static"):
    app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="app/templates")

def _parse_iso_date(s: str):
    try:
        return dt.datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def date_de(s: str) -> str:
    d = _parse_iso_date(s)
    return d.strftime("%d.%m.%Y") if d else (s or "")

def date_wd_de(s: str, style: str = "short") -> str:
    d = _parse_iso_date(s)
    if not d:
        return s or ""
    wd_short = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
    wd_long  = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]
    name = wd_long[d.weekday()] if style == "long" else wd_short[d.weekday()]
    return f"{name}, {d.strftime('%d.%m.%Y')}"

templates.env.filters["date_de"] = date_de
templates.env.filters["date_wd_de"] = date_wd_de

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@app.on_event("startup")
async def startup_event():
    init_db()
    if not ADMIN_TOKEN:
        log.warning("ADMIN_TOKEN not set! Admin UI will be inaccessible.")
    try:
        await refresh_bookings_job()
    except Exception as e:
        log.exception("Initial import failed: %s", e)
    scheduler = AsyncIOScheduler(timezone=TIMEZONE)
    scheduler.add_job(refresh_bookings_job, IntervalTrigger(minutes=REFRESH_INTERVAL_MINUTES))
    scheduler.start()

def _daterange(days=60):
    start = dt.date.today()
    end = start + dt.timedelta(days=days)
    return start.isoformat(), end.isoformat()

def _best_guest_name(it: dict) -> str:
    guest = it.get("guest") or {}
    full = guest.get("fullName") or ""
    if full:
        return full
    fn = guest.get("firstName", "") or it.get("firstName", "")
    ln = guest.get("lastName", "") or it.get("lastName", "")
    name = (fn + " " + ln).strip()
    if name:
        return name
    return it.get("name") or it.get("contactName") or ""

async def refresh_bookings_job():
    client = SmoobuClient()
    start, end = _daterange(60)
    log.info("Refreshing bookings from %s to %s", start, end)
    items = client.get_reservations(start, end)
    log.info("Fetched %d bookings", len(items))
    with SessionLocal() as db:
        seen_booking_ids: List[int] = []
        seen_apartment_ids: List[int] = []
        for it in items:
            b_id = int(it.get("id"))
            apt = it.get("apartment") or {}
            apt_id = int(apt.get("id")) if apt.get("id") is not None else None
            apt_name = apt.get("name") or ""
            guest_name = _best_guest_name(it)

            if apt_id is not None and apt_id not in seen_apartment_ids:
                a = db.get(Apartment, apt_id)
                if not a:
                    a = Apartment(id=apt_id, name=apt_name, planned_minutes=90, active=True)
                    db.add(a)
                else:
                    a.name = apt_name or a.name
                seen_apartment_ids.append(apt_id)

            b = db.get(Booking, b_id)
            if not b:
                b = Booking(id=b_id)
                db.add(b)
            b.apartment_id = apt_id
            b.apartment_name = apt_name or ""
            b.arrival = (it.get("arrival") or "")[:10]
            b.departure = (it.get("departure") or "")[:10]
            b.nights = int(it.get("nights") or 0)
            b.adults = int(it.get("adults") or 1)
            b.children = int(it.get("children") or 0)
            b.guest_comments = (it.get("guestComments") or it.get("comments") or "")[:2000]
            b.guest_name = guest_name or ""
            seen_booking_ids.append(b_id)

        existing_ids = [row[0] for row in db.query(Booking.id).all()]
        for bid in existing_ids:
            if bid not in seen_booking_ids:
                db.delete(db.get(Booking, bid))

        db.commit()

        bookings = db.query(Booking).all()
        upsert_tasks_from_bookings(bookings)

        removed = 0
        for t in db.query(Task).all():
            b = db.get(Booking, t.booking_id) if t.booking_id else None
            if (not t.date or not t.date.strip()) or (not b) or (not b.departure) or (not t.apartment_id):
                db.delete(t); removed += 1
        if removed:
            log.info("Cleanup: %d ungültige Tasks entfernt.", removed)
        db.commit()

@app.get("/", response_class=HTMLResponse)
async def root():
    return "<html><head><link href='https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css' rel='stylesheet'></head><body class='p-4' style='font-family:system-ui;'><h1>Smoobu Staff Planner Pro</h1><p>Service läuft. Admin-UI: <code>/admin/&lt;ADMIN_TOKEN&gt;</code></p><p>Health: <a href='/health'>/health</a></p></body></html>"

@app.get("/health")
async def health():
    return {"ok": True, "time": now_iso()}

# -------------------- Admin UI --------------------
@app.get("/admin/{token}")
async def admin_home(request: Request, token: str, date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None), staff_id: Optional[int] = Query(None), apartment_id: Optional[int] = Query(None), db=Depends(get_db)):
    if token != ADMIN_TOKEN:
        raise HTTPException(status_code=403)
    q = db.query(Task)
    if date_from: q = q.filter(Task.date >= date_from)
    if date_to: q = q.filter(Task.date <= date_to)
    if staff_id: q = q.filter(Task.assigned_staff_id == staff_id)
    if apartment_id: q = q.filter(Task.apartment_id == apartment_id)
    tasks = q.order_by(Task.date, Task.id).all()
    staff = db.query(Staff).filter(Staff.active==True).all()
    apts = db.query(Apartment).filter(Apartment.active==True).all()
    apt_map = {a.id: a.name for a in apts}
    book_map = {b.id: b.guest_name for b in db.query(Booking).all()}
    base_url = BASE_URL.rstrip("/")
    return templates.TemplateResponse("admin_home.html", {"request": request, "token": token, "tasks": tasks, "staff": staff, "apartments": apts, "apt_map": apt_map, "book_map": book_map, "base_url": base_url})

@app.get("/admin/{token}/staff")
async def admin_staff(request: Request, token: str, db=Depends(get_db)):
    if token != ADMIN_TOKEN:
        raise HTTPException(status_code=403)
    staff = db.query(Staff).order_by(Staff.name).all()
    base_url = BASE_URL.rstrip("/")
    return templates.TemplateResponse("admin_staff.html", {"request": request, "token": token, "staff": staff, "base_url": base_url})

@app.post("/admin/{token}/staff/add")
async def admin_staff_add(token: str, name: str = Form(...), hourly_rate: float = Form(0.0), max_hours_per_month: int = Form(160), db=Depends(get_db)):
    if token != ADMIN_TOKEN: raise HTTPException(status_code=403)
    s = Staff(name=name, hourly_rate=hourly_rate, max_hours_per_month=max_hours_per_month, magic_token=new_token(16), active=True)
    db.add(s); db.commit()
    return RedirectResponse(url=f"/admin/{token}/staff", status_code=303)

@app.post("/admin/{token}/staff/toggle")
async def admin_staff_toggle(token: str, staff_id: int = Form(...), db=Depends(get_db)):
    if token != ADMIN_TOKEN: raise HTTPException(status_code=403)
    s = db.get(Staff, staff_id); s.active = not s.active; db.commit()
    return RedirectResponse(url=f"/admin/{token}/staff", status_code=303)

@app.post("/admin/{token}/task/assign")
async def admin_task_assign(token: str, task_id: int = Form(...), staff_id_raw: str = Form(""), db=Depends(get_db)):
    if token != ADMIN_TOKEN: raise HTTPException(status_code=403)
    t = db.get(Task, task_id)
    staff_id: Optional[int] = int(staff_id_raw) if staff_id_raw.strip() else None
    t.assigned_staff_id = staff_id
    db.commit()
    return RedirectResponse(url=f"/admin/{token}", status_code=303)

@app.post("/admin/{token}/task/lock")
async def admin_task_lock(token: str, task_id: int = Form(...), lock: int = Form(1), db=Depends(get_db)):
    if token != ADMIN_TOKEN: raise HTTPException(status_code=403)
    t = db.get(Task, task_id); t.locked = bool(int(lock)); db.commit()
    return RedirectResponse(url=f"/admin/{token}", status_code=303)

@app.get("/admin/{token}/apartments")
async def admin_apartments(request: Request, token: str, db=Depends(get_db)):
    if token != ADMIN_TOKEN:
        raise HTTPException(status_code=403)
    apts = db.query(Apartment).order_by(Apartment.name).all()
    return templates.TemplateResponse("admin_apartments.html", {"request": request, "token": token, "apartments": apts})

@app.post("/admin/{token}/apartments/update")
async def admin_apartments_update(token: str, apartment_id: int = Form(...), planned_minutes: int = Form(...), db=Depends(get_db)):
    if token != ADMIN_TOKEN: raise HTTPException(status_code=403)
    a = db.get(Apartment, apartment_id); a.planned_minutes = int(planned_minutes); db.commit()
    return RedirectResponse(url=f"/admin/{token}/apartments", status_code=303)

@app.get("/admin/{token}/import")
async def admin_import(token: str):
    if token != ADMIN_TOKEN: raise HTTPException(status_code=403)
    await refresh_bookings_job()
    return PlainTextResponse("Import done.")

@app.get("/admin/{token}/timelogs")
async def admin_timelogs(token: str, month: str, db=Depends(get_db)):
    if token != ADMIN_TOKEN: raise HTTPException(status_code=403)
    if len(month) != 7 or month[4] != '-':
        raise HTTPException(400, "month muss Format YYYY-MM haben")
    apts = {a.id: a.name for a in db.query(Apartment).all()}
    staff_map = {s.id: s for s in db.query(Staff).all()}
    rows = []
    for tl in db.query(TimeLog).all():
        if not tl.actual_minutes:
            continue
        if not tl.started_at or tl.started_at[:7] != month:
            continue
        t = db.get(Task, tl.task_id)
        if not t:
            continue
        s = staff_map.get(tl.staff_id)
        rate = float(s.hourly_rate) if s else 0.0
        cost = round((float(tl.actual_minutes)/60.0)*rate, 2)
        rows.append([
            s.name if s else "",
            t.date,
            apts.get(t.apartment_id, ""),
            t.id,
            int(tl.actual_minutes),
            rate,
            cost
        ])
    totals = {}
    for r in rows:
        totals[r[0]] = round(totals.get(r[0], 0.0) + float(r[-1]), 2)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Timelogs {month}"
    headers = ["Mitarbeiter","Datum","Apartment","TaskID","Minuten","Stundensatz (€)","Kosten (€)"]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws2 = wb.create_sheet("Summen")
    ws2.append(["Mitarbeiter","Summe (€)"])
    for name, val in totals.items():
        ws2.append([name, val])
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 18

    bytes_io = io.BytesIO()
    wb.save(bytes_io); bytes_io.seek(0)
    filename = f"timelogs-{month}.xlsx"
    return StreamingResponse(bytes_io, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

# -------------------- Cleaner --------------------
@app.get("/cleaner/{token}")
async def cleaner_home(request: Request, token: str, show_done: int = 0, db=Depends(get_db)):
    s = db.query(Staff).filter(Staff.magic_token==token, Staff.active==True).first()
    if not s: raise HTTPException(status_code=403)
    q = db.query(Task).filter(Task.assigned_staff_id==s.id)
    if not show_done:
        q = q.filter(Task.status != "done")
    tasks = q.order_by(Task.date, Task.id).all()
    apts = db.query(Apartment).all()
    apt_map = {a.id: a.name for a in apts}
    book_map = {b.id: b.guest_name for b in db.query(Booking).all()}
    month = dt.date.today().strftime("%Y-%m")
    planned_total = sum(int(t.planned_minutes or 0) for t in tasks)
    minutes = 0; cost_total = 0.0
    logs = db.query(TimeLog).filter(TimeLog.staff_id==s.id, TimeLog.actual_minutes!=None).all()
    for tl in logs:
        if tl.started_at and tl.started_at[:7]==month and tl.actual_minutes:
            minutes += int(tl.actual_minutes)
            cost_total += (float(tl.actual_minutes)/60.0) * float(s.hourly_rate or 0.0)
    used_hours = round(minutes/60.0, 2)
    cost_total = round(cost_total, 2)
    run_map: Dict[int, str] = {}
    for t in tasks:
        tl = db.query(TimeLog).filter(TimeLog.task_id==t.id, TimeLog.staff_id==s.id, TimeLog.ended_at==None).order_by(TimeLog.id.desc()).first()
        if tl:
            run_map[t.id] = tl.started_at
    warn_limit = used_hours > float(s.max_hours_per_month or 0)
    return templates.TemplateResponse("cleaner.html", {"request": request, "tasks": tasks, "used_hours": used_hours, "apt_map": apt_map, "book_map": book_map, "staff": s, "show_done": show_done, "run_map": run_map, "warn_limit": warn_limit, "planned_total": planned_total, "cost_total": cost_total})

@app.post("/cleaner/{token}/start")
async def cleaner_start(token: str, task_id: int = Form(...), db=Depends(get_db)):
    s = db.query(Staff).filter(Staff.magic_token==token, Staff.active==True).first()
    if not s: raise HTTPException(status_code=403)
    t = db.get(Task, task_id)
    open_tl = db.query(TimeLog).filter(TimeLog.staff_id==s.id, TimeLog.ended_at==None).first()
    if open_tl:
        from datetime import datetime
        open_tl.ended_at = now_iso()
        fmt = "%Y-%m-%d %H:%M:%S"
        open_tl.actual_minutes = int((datetime.strptime(open_tl.ended_at, fmt)-datetime.strptime(open_tl.started_at, fmt)).total_seconds()//60)
    tl = TimeLog(task_id=task_id, staff_id=s.id, started_at=now_iso(), ended_at=None, actual_minutes=None)
    t.status = "running"
    db.add(tl); db.commit()
    return RedirectResponse(url=f"/cleaner/{token}", status_code=303)

@app.post("/cleaner/{token}/stop")
async def cleaner_stop(token: str, task_id: int = Form(...), db=Depends(get_db)):
    s = db.query(Staff).filter(Staff.magic_token==token, Staff.active==True).first()
    if not s: raise HTTPException(status_code=403)
    t = db.get(Task, task_id)
    tl = db.query(TimeLog).filter(TimeLog.task_id==task_id, TimeLog.staff_id==s.id, TimeLog.ended_at==None).first()
    if not tl: raise HTTPException(status_code=404)
    from datetime import datetime
    fmt = "%Y-%m-%d %H:%M:%S"
    tl.ended_at = now_iso()
    start = datetime.strptime(tl.started_at, fmt)
    end = datetime.strptime(tl.ended_at, fmt)
    tl.actual_minutes = int((end-start).total_seconds()//60)
    t.status = "open"
    db.commit()
    return RedirectResponse(url=f"/cleaner/{token}", status_code=303)

@app.post("/cleaner/{token}/done")
async def cleaner_done(token: str, task_id: int = Form(...), db=Depends(get_db)):
    s = db.query(Staff).filter(Staff.magic_token==token, Staff.active==True).first()
    if not s: raise HTTPException(status_code=403)
    t = db.get(Task, task_id)
    t.status = "done"
    db.commit()
    return RedirectResponse(url=f"/cleaner/{token}", status_code=303)

@app.post("/cleaner/{token}/reopen")
async def cleaner_reopen(token: str, task_id: int = Form(...), db=Depends(get_db)):
    s = db.query(Staff).filter(Staff.magic_token==token, Staff.active==True).first()
    if not s: raise HTTPException(status_code=403)
    t = db.get(Task, task_id)
    tl = db.query(TimeLog).filter(TimeLog.task_id==task_id, TimeLog.staff_id==s.id).order_by(TimeLog.id.desc()).first()
    if tl:
        db.delete(tl)
    t.status = "open"
    db.commit()
    return RedirectResponse(url=f"/cleaner/{token}?show_done=1", status_code=303)

@app.post("/cleaner/{token}/note")
async def cleaner_note(token: str, task_id: int = Form(...), note: str = Form(""), db=Depends(get_db)):
    s = db.query(Staff).filter(Staff.magic_token==token, Staff.active==True).first()
    if not s: raise HTTPException(status_code=403)
    t = db.get(Task, task_id)
    t.notes = (note or "").strip()
    db.commit()
    return JSONResponse({"ok": True, "task_id": t.id, "note": t.notes})
